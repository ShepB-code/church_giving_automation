import openpyxl
import json
import datetime
from openpyxl import Workbook

from dotenv import load_dotenv

load_dotenv()

import os

wb = openpyxl.load_workbook(f"{os.environ.get('EXCEL_NAME')}")

sheet = wb.active
def check(row, column):
    if sheet.cell(row=row, column=column).value: 
        return True 
    else:
        return False

def write_json(data):
    with open("giving.json", "w") as outfile: 
        json.dump(data, outfile, indent=4)

def get_name(my_str):
    my_str = my_str.replace("Giving:", "")
    my_str = my_str.replace("Benevolence:", "")
    my_str = my_str.replace("Taxes:", "")
    return my_str

def get_bad_checks(date_col, giving_type_col, amount_col, check_col, year_filter):
    row = 1
    black_listed_checks = list()
    for cell in sheet["F"]: #Category Column 
        if check(row, date_col) and check(row, giving_type_col) and check(row, amount_col):
            if "Giving" in cell.value:
                if year_filter in sheet.cell(row=row, column=date_col).value.strftime("%d-%b-%Y"):
                    if sheet.cell(row=row, column=amount_col).value < 0: 
                        black_listed_checks.append(sheet.cell(row=row, column=check_col).value)
        row += 1

    return black_listed_checks

def store_data(date_col, category_col, amount_col, check_col, year_filter):
    row = 0
    name_dict = dict()
    category = sheet["F"]
    black_listed_checks = get_bad_checks(date_col, category_col, amount_col, check_col, year_filter)

    for cell in category: #Category Column 
        row += 1 #Running count is here because if the loop continues (on a bad check) then the count is off. This way it always executes.
        if check(row, date_col) and check(row, category_col) and check(row, amount_col):
            if "Giving" in cell.value:
                gifter = get_name(cell.value)
                if year_filter in sheet.cell(row=row, column=date_col).value.strftime("%d-%b-%Y"):
                    if sheet.cell(row=row, column=check_col).value in black_listed_checks:
                        continue
                    
                    #Doesn't work yet because not all checks are included.
                    if gifter not in name_dict.keys():
                        name_dict[gifter] = dict()
                        name_dict[gifter]["Total"] = 0
                
                    if "Benevolence" in cell.value:
                        gift_type = "Benevolence"
                    elif "Giving" in cell.value:
                        gift_type = "Giving"
                    
                    if not check(row, check_col):
                        pay_type = "Cash"
                    else:
                        pay_type = sheet.cell(row=row, column=check_col).value

                    date = sheet.cell(row=row, column=date_col).value.strftime("%d-%b-%Y")
                    if gift_type not in name_dict[gifter].keys():
                        name_dict[gifter][gift_type] = dict()
                    amount = sheet.cell(row=row, column=amount_col).value
                    name_dict[gifter][gift_type][date] = [pay_type, amount]
                    name_dict[gifter]["Total"] += amount
                    
        
    return name_dict
